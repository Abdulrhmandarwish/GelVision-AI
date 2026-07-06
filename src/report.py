"""Report generation module for GelVision AI.

Produces downloadable CSV, styled Excel (.xlsx), and professional PDF reports
containing annotated gel images, calibration statistics, and band quantitation
results.
"""

from __future__ import annotations

import datetime
import io
import os
from typing import Optional, Union

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# -----------------------------------------------------------------------
# CSV export
# -----------------------------------------------------------------------


def generate_csv(df: pd.DataFrame) -> bytes:
    """Convert a results DataFrame to CSV bytes.

    Args:
        df: Band quantitation results with columns such as ``Lane``,
            ``Band``, ``Estimated_Size_bp``, ``Intensity``, etc.

    Returns:
        UTF-8 encoded CSV content as raw bytes suitable for download.

    Raises:
        ValueError: If *df* is ``None``.
    """
    if df is None:
        raise ValueError("Cannot generate CSV from a None DataFrame.")
    return df.to_csv(index=False).encode("utf-8")


# -----------------------------------------------------------------------
# Excel export
# -----------------------------------------------------------------------


def generate_excel(df: pd.DataFrame) -> bytes:
    """Convert a results DataFrame to a styled Excel file.

    Styling follows the GelVision AI dark-themed palette:

    * **Header row** — dark navy background (``#0A0F1E``), teal text
      (``#00D4B8``), bold, centred.
    * **Data cells** — centred, thin light-grey borders, auto-width columns.

    Args:
        df: Band quantitation results.

    Returns:
        Raw bytes of the ``.xlsx`` file.

    Raises:
        ValueError: If *df* is ``None``.
    """
    if df is None:
        raise ValueError("Cannot generate Excel from a None DataFrame.")

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Gel Analysis")
        workbook = writer.book
        worksheet = writer.sheets["Gel Analysis"]

        # ── Colour palette ─────────────────────────────────────────
        dark_navy = "0A0F1E"
        teal = "00D4B8"

        header_fill = PatternFill(
            start_color=dark_navy, end_color=dark_navy, fill_type="solid"
        )
        header_font = Font(name="Arial", size=11, bold=True, color=teal)

        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        # ── Style header cells ─────────────────────────────────────
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # ── Style data cells ───────────────────────────────────────
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        # ── Auto-fit column widths ─────────────────────────────────
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    return output.getvalue()


# -----------------------------------------------------------------------
# PDF export (internal builder + public wrapper)
# -----------------------------------------------------------------------


def _build_pdf_report(
    df: pd.DataFrame,
    annotated_img_path: str,
    r2_val: float,
    output_target: Union[str, io.BytesIO],
) -> None:
    """Build a professional PDF report and write it to *output_target*.

    Layout (A4, 36 pt margins):
      1. Title — "GelVision AI — Analysis Report" (Helvetica-Bold 22 pt).
      2. Metadata line — date, lane count, band count.
      3. Annotated gel image (5.5 in × 4.125 in).
      4. Calibration section — R² score with colour-coded quality indicator.
      5. Results table — dark navy headers with teal text, alternating row
         backgrounds.

    Args:
        df: Band quantitation results DataFrame.
        annotated_img_path: Path to the annotated gel image file on disk.
        r2_val: Calibration R² goodness-of-fit score.
        output_target: Either a file path (``str``) or an in-memory
            ``BytesIO`` buffer to write the PDF into.
    """
    doc = SimpleDocTemplate(
        output_target,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()

    # ── Custom paragraph styles ────────────────────────────────────
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=22,
        textColor=colors.HexColor("#0A0F1E"),
        alignment=1,  # centre
        spaceAfter=15,
    )
    h2_style = ParagraphStyle(
        "SectionHeader",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=colors.HexColor("#0A0F1E"),
        spaceBefore=10,
        spaceAfter=6,
        keepWithNext=True,
    )
    body_style = ParagraphStyle(
        "BodyText",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#333333"),
        spaceAfter=8,
    )
    meta_style = ParagraphStyle(
        "MetadataText",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=9,
        textColor=colors.HexColor("#666666"),
        alignment=1,  # centre
    )

    story: list = []

    # ── Title ──────────────────────────────────────────────────────
    story.append(Paragraph("GelVision AI — Analysis Report", title_style))

    # ── Metadata ───────────────────────────────────────────────────
    date_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    n_lanes = int(df["Lane"].nunique()) if not df.empty else 0
    n_bands = len(df)
    meta_html = (
        f"<b>Date:</b> {date_str} &nbsp;|&nbsp; "
        f"<b>Lanes Analyzed:</b> {n_lanes} &nbsp;|&nbsp; "
        f"<b>Bands Found:</b> {n_bands}"
    )
    story.append(Paragraph(meta_html, meta_style))
    story.append(Spacer(1, 10))

    # ── Annotated gel image ────────────────────────────────────────
    story.append(
        Paragraph("Annotated Gel Electrophoresis Image", h2_style)
    )
    if os.path.exists(annotated_img_path):
        story.append(
            Image(annotated_img_path, width=5.5 * inch, height=4.125 * inch)
        )
    else:
        story.append(
            Paragraph(
                "[Annotated image missing from report directory]", body_style
            )
        )
    story.append(Spacer(1, 10))

    # ── Calibration details ────────────────────────────────────────
    story.append(Paragraph("Ladder Calibration Details", h2_style))

    calib_desc = (
        f"Calibration log-linear regression R² fit score: "
        f"<b>{r2_val:.4f}</b>"
    )
    if r2_val >= 0.95:
        calib_desc += (
            " — <font color='#00D4B8'><b>Excellent fit "
            "(R² ≥ 0.95)</b></font>"
        )
    else:
        calib_desc += (
            " — <font color='#FF6B2B'><b>Suboptimal fit "
            "(R² &lt; 0.95)</b></font>"
        )
    story.append(Paragraph(calib_desc, body_style))
    story.append(Spacer(1, 10))

    # ── Results table ──────────────────────────────────────────────
    story.append(Paragraph("Band Quantitation Results", h2_style))

    table_data = [
        ["Lane #", "Band #", "Centroid (y)", "Estimated Size (bp)", "Mean Intensity"]
    ]
    for _, row in df.iterrows():
        table_data.append(
            [
                str(int(row["Lane"])),
                str(int(row["Band"])),
                f"{row['Position']:.1f}",
                f"{int(row['Estimated_Size_bp'])} bp",
                f"{row['Intensity']:.2f}",
            ]
        )

    # Column widths (must sum to ≈ 520 pt for A4 with 36 pt margins)
    col_widths = [80, 80, 100, 130, 130]

    table = Table(table_data, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                # Header row
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0F1E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#00D4B8")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
                ("TOPPADDING", (0, 0), (-1, 0), 5),
                # All cells
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
                # Data rows
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
                ("TOPPADDING", (0, 1), (-1, -1), 4),
                # Alternating row backgrounds
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.HexColor("#F8F9FA"), colors.white],
                ),
            ]
        )
    )

    story.append(table)
    doc.build(story)


def generate_pdf(
    df: pd.DataFrame,
    annotated_img_path: str,
    r2_val: float,
) -> bytes:
    """Generate a PDF analysis report in memory and return raw bytes.

    This is a convenience wrapper around :func:`_build_pdf_report` that
    writes into a ``BytesIO`` buffer instead of a file path.

    Args:
        df: Band quantitation results DataFrame.
        annotated_img_path: Path to the annotated gel image.
        r2_val: Calibration R² goodness-of-fit score.

    Returns:
        Raw PDF file content as bytes, suitable for Streamlit's
        ``st.download_button``.

    Raises:
        ValueError: If *df* is ``None``.
    """
    if df is None:
        raise ValueError("Cannot generate PDF from a None DataFrame.")
    output = io.BytesIO()
    _build_pdf_report(df, annotated_img_path, r2_val, output)
    return output.getvalue()
